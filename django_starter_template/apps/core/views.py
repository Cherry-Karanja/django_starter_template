"""
Core views for the application
"""
from django.http import JsonResponse
from django.middleware.csrf import get_token
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status, viewsets, filters
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiResponse, OpenApiTypes
import csv
import io
from django.db import transaction
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework import generics, permissions, filters
from django.contrib.admin.models import LogEntry
from .serializers import LogEntrySerializer
from celery.result import AsyncResult
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiResponse, OpenApiTypes
import django_filters
from .filters import LogEntryFilter
from .schema import common_responses, pagination_parameters


class BaseModelViewSet(viewsets.ModelViewSet):
    """
    Base ModelViewSet with enhanced features including bulk operations, import/export, and statistics.

    This base viewset provides:
    - Standard CRUD operations from ModelViewSet
    - Search and ordering filters
    - Bulk creation endpoint
    - Bulk import from CSV
    - Bulk export to CSV
    - Statistics endpoint

    Subclasses should define:
    - queryset
    - serializer_class (or get_serializer_class)
    - search_fields (optional)
    - ordering_fields (optional)
    - ordering (optional)
    - permission_classes
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    parser_classes = [MultiPartParser, FormParser]

    @extend_schema(
        summary="Bulk create objects",
        description="Create multiple objects in a single request",
        request={'type': 'array', 'items': {'type': 'object'}},
        responses={
            201: OpenApiResponse(
                description="Objects created successfully",
                response={
                    'type': 'object',
                    'properties': {
                        'count': {'type': 'integer', 'description': 'Number of objects created'},
                        'objects': {'type': 'array', 'items': {'type': 'object'}},
                    }
                }
            ),
            **common_responses
        }
    )
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple objects in bulk"""
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            instances = serializer.save()

        response_serializer = self.get_serializer(instances, many=True)
        return Response({
            'count': len(instances),
            'objects': response_serializer.data
        }, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Bulk import from CSV or Excel",
        description="Import objects from a CSV or Excel (.xlsx) file",
        request={
            'type': 'object',
            'properties': {
                'file': {'type': 'string', 'format': 'binary', 'description': 'CSV or Excel file to import'},
            }
        },
        responses={
            200: OpenApiResponse(
                description="Import completed",
                response={
                    'type': 'object',
                    'properties': {
                        'imported': {'type': 'integer', 'description': 'Number of objects imported'},
                        'errors': {'type': 'array', 'items': {'type': 'string'}},
                    }
                }
            ),
            **common_responses
        }
    )
    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        """Import objects from CSV or Excel file"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        file_name = file.name.lower()

        if file_name.endswith('.csv'):
            # Handle CSV
            file_data = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(file_data))
            rows = list(reader)
        elif file_name.endswith('.xlsx'):
            # Handle Excel
            workbook = load_workbook(file, read_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]  # First row as headers
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                rows.append(row_dict)
        else:
            return Response({'error': 'File must be CSV or Excel (.xlsx)'}, status=status.HTTP_400_BAD_REQUEST)

        imported = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(rows, start=2 if file_name.endswith('.csv') else 2):
                try:
                    serializer = self.get_serializer(data=row)
                    if serializer.is_valid():
                        serializer.save()
                        imported += 1
                    else:
                        errors.append(f"Row {row_num}: {serializer.errors}")
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

        return Response({
            'imported': imported,
            'errors': errors
        })

    @extend_schema(
        summary="Bulk export to CSV or Excel",
        description="Export all objects to CSV or Excel format",
        parameters=[
            OpenApiParameter(name="fields", type=OpenApiTypes.STR, description="Comma-separated list of fields to export"),
            OpenApiParameter(name="format", type=OpenApiTypes.STR, description="Export format: 'csv' or 'xlsx' (default: 'xlsx')", default="xlsx"),
        ],
        responses={
            200: OpenApiResponse(
                description="CSV or Excel file",
                response={'type': 'string', 'format': 'binary'}
            ),
            **common_responses
        }
    )
    @action(detail=False, methods=['get'])
    def bulk_export(self, request):
        """Export objects to CSV or Excel"""
        queryset = self.filter_queryset(self.get_queryset())
        fields = request.query_params.get('fields', None)
        export_format = request.query_params.get('format', 'xlsx').lower()

        if fields:
            fields = [f.strip() for f in fields.split(',')]
        else:
            # Default to model fields
            model = queryset.model
            fields = [f.name for f in model._meta.fields]

        if export_format == 'csv':
            # CSV export
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(fields)

            for obj in queryset:
                row = []
                for field in fields:
                    value = getattr(obj, field, '')
                    if callable(value):
                        value = value()
                    row.append(str(value))
                writer.writerow(row)

            output.seek(0)
            response = Response(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{self.queryset.model._meta.model_name}.csv"'
        elif export_format == 'xlsx':
            # Excel export
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = self.queryset.model._meta.model_name

            # Write headers
            for col_num, field in enumerate(fields, 1):
                sheet.cell(row=1, column=col_num, value=field)

            # Write data
            for row_num, obj in enumerate(queryset, 2):
                for col_num, field in enumerate(fields, 1):
                    value = getattr(obj, field, '')
                    if callable(value):
                        value = value()
                    sheet.cell(row=row_num, column=col_num, value=str(value))

            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            response = Response(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{self.queryset.model._meta.model_name}.xlsx"'
        else:
            return Response({'error': 'Invalid format. Use "csv" or "xlsx"'}, status=status.HTTP_400_BAD_REQUEST)

        return response

    @extend_schema(
        summary="Get statistics",
        description="Get basic statistics for the model",
        responses={
            200: OpenApiResponse(
                description="Model statistics",
                response={
                    'type': 'object',
                    'properties': {
                        'total_count': {'type': 'integer', 'description': 'Total number of objects'},
                        'filtered_count': {'type': 'integer', 'description': 'Number of objects after filtering'},
                    }
                }
            ),
            **common_responses
        }
    )
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get basic statistics"""
        queryset = self.get_queryset()
        filtered_queryset = self.filter_queryset(queryset)

        stats = {
            'total_count': queryset.count(),
            'filtered_count': filtered_queryset.count(),
        }

        # Allow subclasses to add more stats
        extra_stats = self.get_extra_statistics(request, queryset, filtered_queryset)
        stats.update(extra_stats)

        return Response(stats)

    def get_extra_statistics(self, request, queryset, filtered_queryset):
        """
        Override this method in subclasses to add custom statistics.
        Should return a dict of additional statistics.
        """
        return {}


@extend_schema(
    summary="Health check endpoint",
    description="Simple health check to verify the API is running",
    responses={
        200: OpenApiResponse(
            description="API is healthy",
            response={
                'type': 'object',
                'properties': {
                    'status': {'type': 'string', 'example': 'healthy'},
                    'message': {'type': 'string', 'example': 'API is running'},
                }
            }
        ),
        **common_responses
    },
    tags=['Core']
)
@api_view(['GET'])
@permission_classes([AllowAny])
def health_check(request):
    """
    Health check endpoint for monitoring and load balancers.
    """
    return Response({
        'status': 'healthy',
        'message': 'API is running'
    }, status=status.HTTP_200_OK)


@extend_schema(
    summary="Get CSRF token",
    description="Get CSRF token for frontend applications that need session-based authentication",
    responses={
        200: OpenApiResponse(
            description="CSRF token",
            response={
                'type': 'object',
                'properties': {
                    'csrfToken': {'type': 'string', 'description': 'CSRF token for session authentication'},
                }
            }
        ),
        **common_responses
    },
    tags=['Core']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@ensure_csrf_cookie
def get_csrf_token(request):
    """
    Get CSRF token for frontend applications.

    This endpoint ensures a CSRF cookie is set and returns the token
    for frontend applications that need to make session-authenticated requests.
    """
    csrf_token = get_token(request)
    return Response({
        'csrfToken': csrf_token
    }, status=status.HTTP_200_OK)


@require_http_methods(["GET"])
@ensure_csrf_cookie
def csrf_token_view(request):
    """
    Alternative CSRF token endpoint for non-DRF clients.
    Returns CSRF token as JSON.
    """
    csrf_token = get_token(request)
    return JsonResponse({'csrfToken': csrf_token})


@extend_schema(
    summary="List history items (audit log)",
    description="Returns a paginated list of history items (audit log entries) for all models in the system. Only staff/admins can view.",
    parameters=[
        OpenApiParameter(name="user", location=OpenApiParameter.QUERY, type=OpenApiTypes.STR, description="Filter by username (icontains)"),
        OpenApiParameter(name="user_id", location=OpenApiParameter.QUERY, type=OpenApiTypes.INT, description="Filter by user ID"),
        OpenApiParameter(name="action_time_after", location=OpenApiParameter.QUERY, type=OpenApiTypes.DATETIME, description="Filter by action time after"),
        OpenApiParameter(name="action_time_before", location=OpenApiParameter.QUERY, type=OpenApiTypes.DATETIME, description="Filter by action time before"),
        OpenApiParameter(name="action_flag", location=OpenApiParameter.QUERY, type=OpenApiTypes.INT, description="Filter by action flag (1=Addition, 2=Change, 3=Deletion)"),
        OpenApiParameter(name="content_type", location=OpenApiParameter.QUERY, type=OpenApiTypes.STR, description="Filter by model name (icontains)"),
        OpenApiParameter(name="object_id", location=OpenApiParameter.QUERY, type=OpenApiTypes.STR, description="Filter by object ID (icontains)"),
        OpenApiParameter(name="object_repr", location=OpenApiParameter.QUERY, type=OpenApiTypes.STR, description="Filter by object representation (icontains)"),
        OpenApiParameter(name="change_message", location=OpenApiParameter.QUERY, type=OpenApiTypes.STR, description="Filter by change message (icontains)"),
    ] + pagination_parameters,
    responses={
        200: LogEntrySerializer(many=True),
        **common_responses
    },
    tags=["Core"]
)
class HistoryListView(generics.ListAPIView):
    queryset = LogEntry.objects.all().order_by('-action_time')
    serializer_class = LogEntrySerializer
    filter_backends = [django_filters.rest_framework.DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = LogEntryFilter
    search_fields = ['user__username', 'object_repr', 'change_message', 'content_type__model']
    ordering_fields = ['action_time', 'user__username', 'content_type__model', 'action_flag']
    ordering = ['-action_time']
    permission_classes = [permissions.IsAdminUser]


@extend_schema(
    summary="Get task status",
    description="Check the status of an asynchronous Celery task by its task ID",
    parameters=[
        OpenApiParameter(
            name="task_id",
            location=OpenApiParameter.PATH,
            type=OpenApiTypes.UUID,
            description="The UUID of the Celery task to check",
            required=True
        )
    ],
    responses={
        200: OpenApiResponse(
            description="Task status information",
            response={
                'type': 'object',
                'properties': {
                    'task_id': {'type': 'string', 'format': 'uuid', 'description': 'The task ID'},
                    'status': {'type': 'string', 'enum': ['PENDING', 'PROGRESS', 'SUCCESS', 'FAILURE', 'RETRY', 'REVOKED'], 'description': 'Current task status'},
                    'result': {'type': 'object', 'description': 'Task result (only present when status is SUCCESS)'},
                    'error': {'type': 'string', 'description': 'Error message (only present when status is FAILURE)'},
                    'traceback': {'type': 'string', 'description': 'Error traceback (only present when status is FAILURE)'},
                    'date_done': {'type': 'string', 'format': 'date-time', 'description': 'When the task completed (only present when task is done)'},
                }
            }
        ),
        **common_responses
    },
    tags=['Core']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def task_status(request, task_id):
    """
    Get the status of a Celery task by its ID.

    This endpoint allows clients to check the progress and result of
    asynchronous tasks like AI chat responses, document generation, etc.
    """
    # task_id is already validated as UUID by Django's URL converter
    # No need to validate again since <uuid:task_id> ensures it's a valid UUID

    # Get task result from Celery
    task_result = AsyncResult(str(task_id))

    if task_result.state == "PENDING":
        # Task is waiting to be processed
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
        }
    elif task_result.state == "PROGRESS":
        # Task is in progress (if using progress tracking)
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
            "current": getattr(task_result.info, 'current', None),
            "total": getattr(task_result.info, 'total', None),
            "message": getattr(task_result.info, 'message', ''),
        }
    elif task_result.state == "SUCCESS":
        # Task completed successfully
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
            "result": task_result.result,
            "date_done": task_result.date_done.isoformat() if task_result.date_done else None,
        }
    elif task_result.state == "FAILURE":
        # Task failed
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
            "error": str(task_result.info),
            "traceback": getattr(task_result.info, 'traceback', None),
            "date_done": task_result.date_done.isoformat() if task_result.date_done else None,
        }
    else:
        # Other states (RETRY, REVOKED, etc.)
        response_data = {
            "task_id": task_id,
            "status": task_result.state,
            "info": str(task_result.info) if task_result.info else None,
            "date_done": task_result.date_done.isoformat() if task_result.date_done else None,
        }

    return Response(response_data, status=status.HTTP_200_OK)


@extend_schema(
    summary="Dashboard Statistics",
    description="""
    Get basic dashboard statistics for the application.

    This is a template endpoint that can be customized based on your application's needs.
    Currently returns basic system statistics that can be extended.
    """,
    responses={
        200: OpenApiResponse(
            description="Dashboard statistics",
            response={
                'type': 'object',
                'properties': {
                    'total_users': {'type': 'integer', 'description': 'Total number of users'},
                    'active_users': {'type': 'integer', 'description': 'Number of active users'},
                    'total_log_entries': {'type': 'integer', 'description': 'Total number of audit log entries'},
                    'recent_log_entries': {'type': 'integer', 'description': 'Number of log entries in last 24 hours'},
                }
            }
        ),
        **common_responses
    },
    tags=['Core']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_statistics(request):
    """
    Get basic dashboard statistics for the application.

    This endpoint provides a template for dashboard statistics.
    Customize this based on your application's specific models and requirements.
    """
    from accounts.models import User
    from django.utils import timezone
    from datetime import timedelta

    # Basic user statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()

    # Audit log statistics
    total_log_entries = LogEntry.objects.count()
    yesterday = timezone.now() - timedelta(days=1)
    recent_log_entries = LogEntry.objects.filter(action_time__gte=yesterday).count()

    statistics = {
        'total_users': total_users,
        'active_users': active_users,
        'total_log_entries': total_log_entries,
        'recent_log_entries': recent_log_entries,
    }

    return Response(statistics)